"""
Automatic result checker for the football betting bot.

Usage:
    python auto_results.py              # run once immediately and exit
    python auto_results.py --schedule   # nightly daemon at 00:15 Brussels
    python auto_results.py --live       # check every 30 min + Telegram alerts
"""
from __future__ import annotations

import logging
import os
import re
import sys
from datetime import date, datetime, timedelta

import requests
from apscheduler.schedulers.blocking import BlockingScheduler
from dotenv import load_dotenv
from openpyxl import load_workbook

from excel_tracker import (
    EXCEL_PATH,
    _style_picks_row,
    finalize_workbook,
    init_excel,
)

load_dotenv()
logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
log = logging.getLogger(__name__)

HOST          = "free-api-live-football-data.p.rapidapi.com"
LOOKBACK_DAYS = 7


# ── Telegram ──────────────────────────────────────────────────────────────────

def _telegram_send(text: str) -> None:
    token   = os.getenv("TELEGRAM_BOT_TOKEN")
    chat_id = os.getenv("TELEGRAM_CHANNEL_ID")
    if not token or not chat_id:
        log.warning("Telegram not configured — skipping notification")
        return
    try:
        r = requests.post(
            f"https://api.telegram.org/bot{token}/sendMessage",
            json={"chat_id": chat_id, "text": text},
            timeout=10,
        )
        r.raise_for_status()
        log.info("Telegram notification sent")
    except Exception as exc:
        log.error("Telegram send failed: %s", exc)


def _score_description(
    bet_type: str,
    pick: str,
    home_name: str,
    away_name: str,
    home_score: int,
    away_score: int,
) -> str:
    """Human-readable result line for a Telegram notification."""
    bt    = bet_type.lower()
    total = home_score + away_score
    score = f"{home_score}-{away_score}"

    if any(x in bt for x in ("over", "under", "total goals", "o/u")):
        return f"{score} ({total} goals total)"

    if any(x in bt for x in ("both teams to score", "btts")):
        if home_score > 0 and away_score > 0:
            suffix = "both teams scored"
        elif home_score == 0 and away_score == 0:
            suffix = "goalless draw"
        else:
            loser = home_name if away_score > home_score else away_name
            suffix = f"{loser} kept a clean sheet"
        return f"{score} ({suffix})"

    if home_score > away_score:
        base = f"{home_name} won {score}"
    elif away_score > home_score:
        base = f"{away_name} won {score}"
    else:
        base = f"{score} draw"

    if any(x in bt for x in ("asian handicap", "handicap")):
        m = re.search(r'([+-]?\d+\.?\d*)\s*$', pick.strip())
        if m:
            hc      = float(m.group(1))
            team_q  = pick[:m.start()].strip().lower()
            if team_q in home_name.lower():
                base += f" (handicap adjusted: {home_score + hc:.1f}-{away_score})"
            elif team_q in away_name.lower():
                base += f" (handicap adjusted: {home_score}-{away_score + hc:.1f})"

    return base


def _format_result_notification(r: dict) -> str:
    emoji   = "✅" if r["result"] == "WIN" else ("❌" if r["result"] == "LOSS" else "⬜")
    pnl_str = f"+{r['pnl']:.2f}" if r["pnl"] >= 0 else f"{r['pnl']:.2f}"
    desc    = _score_description(
        r["bet_type"], r["pick"],
        r["home_name"], r["away_name"],
        r["home_score"], r["away_score"],
    )
    return (
        f"{emoji} {r['result']} — {r['match']}\n"
        f"Bet: {r['bet_type']} | Odds: {r['odds']:.2f}\n"
        f"Pick: {r['pick']}\n"
        f"Result: {desc}\n"
        f"P&L: {pnl_str} units"
    )


# ── API ───────────────────────────────────────────────────────────────────────

def _fetch_matches(dt: date) -> list[dict]:
    headers = {"x-rapidapi-host": HOST, "x-rapidapi-key": os.environ.get("RAPIDAPI_KEY")}
    r = requests.get(
        f"https://{HOST}/football-get-matches-by-date",
        headers=headers,
        params={"date": dt.strftime("%Y%m%d")},
        timeout=15,
    )
    r.raise_for_status()
    return r.json().get("response", {}).get("matches", [])


def _find_api_match(matches: list[dict], home_q: str, away_q: str) -> dict | None:
    hq = home_q.lower().strip()
    aq = away_q.lower().strip()
    for m in matches:
        h = m["home"]["longName"].lower()
        a = m["away"]["longName"].lower()
        if (hq in h or h in hq) and (aq in a or a in aq):
            return m
    return None


# ── Bet-type evaluation ───────────────────────────────────────────────────────

def _parse_handicap(pick: str) -> tuple[str, float] | None:
    m = re.search(r'([+-]?\d+\.?\d*)\s*$', pick.strip())
    if not m:
        return None
    return pick[:m.start()].strip().lower(), float(m.group(1))


def evaluate_pick(
    bet_type: str,
    pick: str,
    home_name: str,
    away_name: str,
    home_score: int,
    away_score: int,
) -> str:
    """
    Return WIN, LOSS, VOID, or PENDING (unrecognised bet type / data missing).

    Handles both generic terms ('Home Win', 'Away or Draw') and team-name
    picks generated by the updated Claude prompt ('Sweden Win', 'Ivory Coast or Draw').
    """
    bt    = bet_type.lower()
    pk    = pick.lower().strip()
    hn    = home_name.lower()
    an    = away_name.lower()
    total = home_score + away_score
    hw    = home_score > away_score
    aw    = away_score > home_score
    dr    = home_score == away_score

    # ── Match Winner ─────────────────────────────────────────────────────────
    if any(x in bt for x in ("match winner", "1x2", "result", "moneyline")):
        # Generic terms OR team name anywhere in pick string
        home_pick = pk in ("home", "home win", "1") or hn in pk or pk in hn
        away_pick = pk in ("away", "away win", "2") or an in pk or pk in an
        draw_pick = pk in ("draw", "x", "tie")

        if home_pick and not away_pick: return "WIN" if hw else "LOSS"
        if away_pick and not home_pick: return "WIN" if aw else "LOSS"
        if draw_pick:                   return "WIN" if dr else "LOSS"

    # ── Both Teams to Score ──────────────────────────────────────────────────
    elif any(x in bt for x in ("both teams to score", "btts", "gg/ng", "goal goal")):
        both = home_score > 0 and away_score > 0
        if pk in ("yes", "true", "gg", "yes (gg)"): return "WIN" if both     else "LOSS"
        if pk in ("no",  "false", "ng", "no (ng)"): return "WIN" if not both else "LOSS"

    # ── Over / Under goals ───────────────────────────────────────────────────
    elif any(x in bt for x in ("over", "under", "total goals", "goals over", "o/u")):
        nums      = re.findall(r'\d+\.?\d*', bt)
        threshold = float(nums[0]) if nums else 2.5
        if "over"  in pk: return "WIN" if total >  threshold else "LOSS"
        if "under" in pk: return "WIN" if total <  threshold else "LOSS"

    # ── Asian Handicap ───────────────────────────────────────────────────────
    elif any(x in bt for x in ("asian handicap", "handicap", " ah ")):
        parsed = _parse_handicap(pick)
        if parsed:
            team_q, hc = parsed
            if team_q in hn or hn in team_q:
                adj = home_score + hc
                if   adj > away_score: return "WIN"
                elif adj < away_score: return "LOSS"
                else:                  return "VOID"
            elif team_q in an or an in team_q:
                adj = away_score + hc
                if   adj > home_score: return "WIN"
                elif adj < home_score: return "LOSS"
                else:                  return "VOID"

    # ── Double Chance ────────────────────────────────────────────────────────
    elif "double chance" in bt:
        # "or draw" picks: home team name or "home" must appear alongside "or draw"
        if "or draw" in pk:
            home_side = hn in pk or "home" in pk or "1x" in pk
            away_side = an in pk or "away" in pk or "x2" in pk
            if home_side and not away_side: return "WIN" if hw or dr else "LOSS"
            if away_side and not home_side: return "WIN" if aw or dr else "LOSS"
        # Both-wins double chance
        if any(x in pk for x in ("home or away", "12")) or (hn in pk and an in pk):
            return "WIN" if hw or aw else "LOSS"

    log.warning("evaluate_pick: unhandled bet_type='%s' pick='%s'", bet_type, pick)
    return "PENDING"


# ── Core checker ─────────────────────────────────────────────────────────────

def run_auto_results(lookback_days: int = LOOKBACK_DAYS) -> tuple[dict, list[dict]]:
    """
    Scan pending Excel rows, fetch API scores, update the workbook.
    Returns (stats_dict, list_of_newly_resolved_picks).
    """
    init_excel()
    if not EXCEL_PATH.exists():
        log.info("picks_tracker.xlsx not found — nothing to check.")
        return {}, []

    wb = load_workbook(EXCEL_PATH)
    ws = wb["Picks"]

    cutoff   = date.today() - timedelta(days=lookback_days)
    stats    = {"checked": 0, "updated": 0, "not_finished": 0,
                "no_match": 0, "too_old": 0, "errors": 0}
    changed  = False
    resolved: list[dict] = []

    # ── 1. Collect pending rows ───────────────────────────────────────────────
    pending: list[dict] = []
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=1).value is None:
            break
        if ws.cell(row=row, column=7).value:
            continue

        dt_val    = ws.cell(row=row, column=1).value
        pick_date = dt_val.date() if isinstance(dt_val, datetime) else dt_val

        if pick_date < cutoff:
            stats["too_old"] += 1
            continue

        pending.append({
            "row":      row,
            "date":     pick_date,
            "match":    ws.cell(row=row, column=2).value or "",
            "bet_type": ws.cell(row=row, column=3).value or "",
            "pick":     ws.cell(row=row, column=4).value or "",
            "odds":     float(ws.cell(row=row, column=5).value or 1.0),
        })

    if not pending:
        log.info("No pending picks in the lookback window.")
        return stats, resolved

    log.info("Found %d pending pick(s) to check.", len(pending))

    # ── 2. Batch API calls by date ────────────────────────────────────────────
    api_cache: dict[date, list[dict]] = {}
    for p in pending:
        for dt in (p["date"], p["date"] + timedelta(days=1)):
            if dt in api_cache:
                continue
            try:
                api_cache[dt] = _fetch_matches(dt)
                log.info("  API: fetched %d matches for %s", len(api_cache[dt]), dt)
            except Exception as exc:
                log.error("  API fetch failed for %s: %s", dt, exc)
                api_cache[dt] = []

    # ── 3. Evaluate and update ────────────────────────────────────────────────
    for p in pending:
        stats["checked"] += 1
        row      = p["row"]
        match    = p["match"]
        bet_type = p["bet_type"]
        pick     = p["pick"]
        odds     = p["odds"]

        if " vs " not in match:
            log.warning("Row %d: cannot parse match name '%s'", row, match)
            stats["errors"] += 1
            continue

        home_q, away_q = [s.strip() for s in match.split(" vs ", 1)]

        api_match = None
        for dt in (p["date"], p["date"] + timedelta(days=1)):
            api_match = _find_api_match(api_cache.get(dt, []), home_q, away_q)
            if api_match:
                break

        if api_match is None:
            log.info("Row %d: '%s' — not found in API yet", row, match)
            stats["no_match"] += 1
            continue

        if not api_match["status"].get("finished"):
            log.info("Row %d: '%s' — match not finished yet", row, match)
            stats["not_finished"] += 1
            continue

        home_score = int(api_match["home"].get("score") or 0)
        away_score = int(api_match["away"].get("score") or 0)
        home_name  = api_match["home"]["longName"]
        away_name  = api_match["away"]["longName"]

        result = evaluate_pick(bet_type, pick, home_name, away_name, home_score, away_score)

        if result == "PENDING":
            log.warning("Row %d: could not evaluate bet_type='%s' pick='%s'", row, bet_type, pick)
            stats["errors"] += 1
            continue

        pnl = round(odds - 1, 2) if result == "WIN" else (-1.0 if result == "LOSS" else 0.0)
        ws.cell(row=row, column=7).value = result
        ws.cell(row=row, column=8).value = pnl
        _style_picks_row(ws, row, result)
        changed = True
        stats["updated"] += 1

        log.info(
            "Row %d: %s [%s→%s]  score %d-%d  P&L %+.2f",
            row, match, pick, result, home_score, away_score, pnl,
        )

        resolved.append({
            "match":      match,
            "bet_type":   bet_type,
            "pick":       pick,
            "odds":       odds,
            "result":     result,
            "pnl":        pnl,
            "home_name":  home_name,
            "away_name":  away_name,
            "home_score": home_score,
            "away_score": away_score,
        })

    # ── 4. Recalculate and save ───────────────────────────────────────────────
    if changed:
        finalize_workbook(wb)
        wb.save(EXCEL_PATH)
        log.info("Excel saved — %d row(s) updated.", stats["updated"])
    else:
        log.info("No changes — Excel unchanged.")

    return stats, resolved


# ── Entry point ───────────────────────────────────────────────────────────────

def _print_stats(stats: dict) -> None:
    print(f"\n  Checked     : {stats.get('checked', 0)}")
    print(f"  Updated     : {stats.get('updated', 0)}")
    print(f"  Not finished: {stats.get('not_finished', 0)}")
    print(f"  No API match: {stats.get('no_match', 0)}")
    print(f"  Too old     : {stats.get('too_old', 0)}")
    print(f"  Errors      : {stats.get('errors', 0)}")


if __name__ == "__main__":
    live_mode     = "--live"     in sys.argv
    schedule_mode = "--schedule" in sys.argv

    if live_mode:
        # Track which picks have already been notified this session
        notified: set[tuple] = set()

        def _live_check() -> None:
            print("\n--- Checking results ---")
            stats, resolved = run_auto_results(lookback_days=2)
            _print_stats(stats)
            for r in resolved:
                key = (r["match"], r["bet_type"], r["pick"])
                if key in notified:
                    continue
                msg = _format_result_notification(r)
                print(f"\nSending notification:\n{msg}")
                _telegram_send(msg)
                notified.add(key)

        _live_check()

        scheduler = BlockingScheduler()
        scheduler.add_job(_live_check, "interval", minutes=30)
        print("\nLive result checker running — checks every 30 minutes.")
        print("Press Ctrl+C to stop.")
        try:
            scheduler.start()
        except (KeyboardInterrupt, SystemExit):
            print("Live checker stopped.")

    elif schedule_mode:
        scheduler = BlockingScheduler(timezone="Europe/Brussels")
        scheduler.add_job(
            lambda: run_auto_results(),
            "cron",
            hour=0,
            minute=15,
        )
        print("Auto-result checker started — runs nightly at 00:15 Brussels.")
        print("Press Ctrl+C to stop.")
        try:
            scheduler.start()
        except (KeyboardInterrupt, SystemExit):
            print("Scheduler stopped.")

    else:
        print("Running auto-result check now...")
        stats, _ = run_auto_results()
        _print_stats(stats)
        print(f"\nDone.  Excel: {EXCEL_PATH}")
