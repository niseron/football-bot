"""
Excel tracking layer for the football betting bot.
Writes picks_tracker.xlsx alongside the SQLite picks.db.
"""
from __future__ import annotations

import logging
from collections import defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

log = logging.getLogger(__name__)

EXCEL_PATH = Path(__file__).parent / "picks_tracker.xlsx"

PICKS_HEADERS = [
    "Date", "Match", "Bet Type", "Pick", "Odds",
    "Confidence", "Result", "Profit/Loss", "Running Total P&L",
]
COL_WIDTHS = [14, 36, 24, 20, 8, 13, 10, 14, 20]

# ── Styles ────────────────────────────────────────────────────────────────────
_HDR_FILL    = PatternFill("solid", fgColor="1F4E79")
_ALT_FILL    = PatternFill("solid", fgColor="D6E4F0")
_WIN_FILL    = PatternFill("solid", fgColor="C6EFCE")
_LOSS_FILL   = PatternFill("solid", fgColor="FFC7CE")
_VOID_FILL   = PatternFill("solid", fgColor="FFEB9C")
_SUM_FILL    = PatternFill("solid", fgColor="2E75B6")
_SUM_SEC     = PatternFill("solid", fgColor="BDD7EE")
_HDR_FONT    = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
_BODY_FONT   = Font(name="Calibri", size=11)
_BOLD_FONT   = Font(bold=True, name="Calibri", size=11)
_SUM_HDR_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=12)

_PNL_FMT  = '+0.00;-0.00;"0.00"'
_DATE_FMT = "DD-MMM-YYYY"


# ── Workbook bootstrap ────────────────────────────────────────────────────────

def _apply_header(ws, n_cols: int) -> None:
    for col in range(1, n_cols + 1):
        c = ws.cell(row=1, column=col)
        c.font = _HDR_FONT
        c.fill = _HDR_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 22


def _apply_picks_cols(ws) -> None:
    ws.freeze_panes = "A2"
    for i, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.auto_filter.ref = f"A1:{get_column_letter(len(PICKS_HEADERS))}1"


def _style_picks_row(ws, row: int, result: str | None) -> None:
    fill = (
        _WIN_FILL  if result == "WIN"  else
        _LOSS_FILL if result == "LOSS" else
        _VOID_FILL if result == "VOID" else
        _ALT_FILL  if row % 2 == 0    else None
    )
    for col in range(1, len(PICKS_HEADERS) + 1):
        c = ws.cell(row=row, column=col)
        c.font = _BODY_FONT
        if fill:
            c.fill = fill
        if col == 1:
            c.number_format = _DATE_FMT
            c.alignment = Alignment(horizontal="center")
        elif col in (5, 8, 9):
            c.number_format = _PNL_FMT
            c.alignment = Alignment(horizontal="center")
        elif col == 7:
            c.alignment = Alignment(horizontal="center")


def init_excel() -> None:
    if EXCEL_PATH.exists():
        return
    wb = Workbook()
    wb.remove(wb.active)

    ws_p = wb.create_sheet("Picks")
    for i, h in enumerate(PICKS_HEADERS, 1):
        ws_p.cell(row=1, column=i, value=h)
    _apply_header(ws_p, len(PICKS_HEADERS))
    _apply_picks_cols(ws_p)

    wb.create_sheet("Summary")
    wb.save(EXCEL_PATH)
    log.info("Created %s", EXCEL_PATH)


# ── Write a new pick ──────────────────────────────────────────────────────────

def log_to_excel(
    match: str,
    league: str,
    bet_type: str,
    pick: str,
    odds: float,
    confidence: str,
    pick_date: str | None = None,
) -> None:
    init_excel()
    wb = load_workbook(EXCEL_PATH)
    ws = wb["Picks"]

    dt = datetime.fromisoformat(pick_date) if pick_date else datetime.now()
    target_date = dt.date()

    for r in range(2, ws.max_row + 1):
        existing_dt = ws.cell(row=r, column=1).value
        if existing_dt is None:
            break
        existing_date = existing_dt.date() if isinstance(existing_dt, datetime) else None
        if (
            existing_date == target_date
            and ws.cell(row=r, column=2).value == match
            and ws.cell(row=r, column=3).value == bet_type
            and ws.cell(row=r, column=4).value == pick
        ):
            log.info("Excel: skipping duplicate '%s — %s'", match, pick)
            return

    row = ws.max_row + 1

    ws.cell(row=row, column=1, value=dt)
    ws.cell(row=row, column=2, value=match)
    ws.cell(row=row, column=3, value=bet_type)
    ws.cell(row=row, column=4, value=pick)
    ws.cell(row=row, column=5, value=round(float(odds), 2))
    ws.cell(row=row, column=6, value=confidence)
    # columns 7-9 (Result, P/L, Running) stay blank until result is entered

    _style_picks_row(ws, row, None)
    wb.save(EXCEL_PATH)
    log.info("Excel: logged '%s — %s'", match, pick)


# ── Running total + summary refresh ──────────────────────────────────────────

def _recalculate_running_total(ws) -> None:
    running = 0.0
    for row in range(2, ws.max_row + 1):
        result = ws.cell(row=row, column=7).value
        pnl    = ws.cell(row=row, column=8).value
        rt_c   = ws.cell(row=row, column=9)

        if result in ("WIN", "LOSS", "VOID") and pnl is not None:
            running += pnl
            rt_c.value = round(running, 2)
            rt_c.number_format = _PNL_FMT
            rt_c.alignment = Alignment(horizontal="center")
            ws.cell(row=row, column=8).number_format = _PNL_FMT
        else:
            rt_c.value = None


def _refresh_summary(wb: Workbook) -> None:
    ws_picks = wb["Picks"]
    ws_sum   = wb["Summary"]

    # Collect all pick rows
    rows: list[dict] = []
    for r in ws_picks.iter_rows(min_row=2, values_only=True):
        if r[0] is None:
            break
        rows.append({
            "match": r[1], "bet_type": r[2], "pick": r[3],
            "odds": r[4], "confidence": r[5],
            "result": r[6], "pnl": r[7],
        })

    settled = [r for r in rows if r["result"] in ("WIN", "LOSS", "VOID")]
    wins    = [r for r in settled if r["result"] == "WIN"]
    losses  = [r for r in settled if r["result"] == "LOSS"]
    voids   = [r for r in settled if r["result"] == "VOID"]
    pending = [r for r in rows if not r["result"]]

    total_pnl = round(sum(r["pnl"] or 0 for r in settled if r["pnl"] is not None), 2)
    win_rate  = round(len(wins) / len(settled) * 100, 1) if settled else 0.0

    # Best bet type
    bt_pnl: dict[str, float] = defaultdict(float)
    for r in settled:
        if r["pnl"] is not None:
            bt_pnl[r["bet_type"]] += r["pnl"]
    best_bt     = max(bt_pnl, key=bt_pnl.get) if bt_pnl else "N/A"
    best_bt_pnl = round(bt_pnl.get(best_bt, 0), 2)

    # Best confidence level
    conf_pnl: dict[str, float] = defaultdict(float)
    for r in settled:
        if r["pnl"] is not None:
            conf_pnl[r["confidence"]] += r["pnl"]
    best_conf     = max(conf_pnl, key=conf_pnl.get) if conf_pnl else "N/A"
    best_conf_pnl = round(conf_pnl.get(best_conf, 0), 2)

    # Clear and rewrite Summary sheet
    ws_sum.delete_rows(1, ws_sum.max_row + 1)
    ws_sum.column_dimensions["A"].width = 32
    ws_sum.column_dimensions["B"].width = 18

    def _row(label: str, value, bold_label: bool = False,
             section_hdr: bool = False, sub: bool = False):
        return (label, value, bold_label, section_hdr, sub)

    sections = [
        _row("PERFORMANCE SUMMARY", "",         section_hdr=True),
        _row("",                    ""),
        _row("Total picks",         len(rows),  bold_label=True),
        _row("  Wins",              len(wins)),
        _row("  Losses",            len(losses)),
        _row("  Voids",             len(voids)),
        _row("  Pending",           len(pending)),
        _row("",                    ""),
        _row("Win rate",            f"{win_rate}%",    bold_label=True),
        _row("Total P&L (units)",   total_pnl,         bold_label=True),
        _row("",                    ""),
        _row("Best bet type",       best_bt,           bold_label=True),
        _row("  P&L from this type",best_bt_pnl,       sub=True),
        _row("",                    ""),
        _row("Best confidence level", best_conf,       bold_label=True),
        _row("  P&L from this level", best_conf_pnl,  sub=True),
    ]

    for ri, (label, value, bold_label, section_hdr, sub) in enumerate(sections, 1):
        ca = ws_sum.cell(row=ri, column=1, value=label)
        cb = ws_sum.cell(row=ri, column=2, value=value)

        if section_hdr:
            ca.font = _SUM_HDR_FONT
            for col in (1, 2):
                ws_sum.cell(row=ri, column=col).fill = _SUM_FILL
            ws_sum.row_dimensions[ri].height = 22
        elif bold_label:
            ca.font = _BOLD_FONT
            cb.font = _BODY_FONT
        elif sub:
            ca.font = Font(name="Calibri", size=11, color="595959")
            cb.font = Font(name="Calibri", size=11, color="595959")
        else:
            ca.font = _BODY_FONT
            cb.font = _BODY_FONT

        if isinstance(value, float):
            cb.number_format = _PNL_FMT


# ── Public helper used by auto_results ───────────────────────────────────────

def finalize_workbook(wb: Workbook) -> None:
    """Recalculate running totals and refresh Summary sheet, then the caller saves."""
    _recalculate_running_total(wb["Picks"])
    _refresh_summary(wb)


# ── Update a result ───────────────────────────────────────────────────────────

def update_result(match_query: str, pick_query: str, result: str) -> bool:
    """
    Find the most recent PENDING pick where match and pick/bet_type
    match the given queries (case-insensitive, partial). Returns True on success.
    """
    result = result.upper()
    if result not in ("WIN", "LOSS", "VOID"):
        raise ValueError(f"Result must be WIN, LOSS or VOID — got '{result}'")

    init_excel()
    wb = load_workbook(EXCEL_PATH)
    ws = wb["Picks"]

    mq = match_query.lower().strip()
    pq = pick_query.lower().strip()

    target_row = None
    for row in range(ws.max_row, 1, -1):   # bottom-up → most recent first
        m_val  = (ws.cell(row=row, column=2).value or "").lower()
        bt_val = (ws.cell(row=row, column=3).value or "").lower()
        p_val  = (ws.cell(row=row, column=4).value or "").lower()
        res    =  ws.cell(row=row, column=7).value

        match_ok = mq in m_val or m_val in mq
        pick_ok  = (pq in p_val or p_val in pq or
                    pq in bt_val or bt_val in pq or
                    pq in f"{bt_val} {p_val}")

        if match_ok and pick_ok and not res:
            target_row = row
            break

    if target_row is None:
        print(f"No pending pick found matching '{match_query}' / '{pick_query}'")
        return False

    odds = float(ws.cell(row=target_row, column=5).value or 1.0)
    pnl  = round(odds - 1, 2) if result == "WIN" else (-1.0 if result == "LOSS" else 0.0)

    ws.cell(row=target_row, column=7).value = result
    ws.cell(row=target_row, column=8).value = pnl
    _style_picks_row(ws, target_row, result)
    _recalculate_running_total(ws)
    _refresh_summary(wb)

    wb.save(EXCEL_PATH)

    match_name = ws.cell(row=target_row, column=2).value
    sign = "+" if pnl >= 0 else ""
    print(f"Updated  : {match_name}")
    print(f"Result   : {result}")
    print(f"P&L      : {sign}{pnl:.2f} units  (odds {odds})")
    return True


# ── Weekly data extraction ────────────────────────────────────────────────────

def get_weekly_data() -> dict:
    """Return this week's (Mon–Sun UTC) stats from the Excel file."""
    if not EXCEL_PATH.exists():
        return {}

    today    = date.today()
    week_mon = today - timedelta(days=today.weekday())
    week_sun = week_mon + timedelta(days=6)

    wb = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    ws = wb["Picks"]

    all_rows: list[dict] = []
    week_rows: list[dict] = []

    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[0] is None:
            break
        dt = r[0].date() if isinstance(r[0], datetime) else (r[0] or date.today())
        row = {
            "date": dt, "match": r[1], "bet_type": r[2], "pick": r[3],
            "odds": r[4], "confidence": r[5], "result": r[6],
            "pnl": r[7], "running": r[8],
        }
        all_rows.append(row)
        if week_mon <= dt <= week_sun:
            week_rows.append(row)

    wb.close()

    settled = [r for r in week_rows if r["result"] in ("WIN", "LOSS", "VOID")]
    wins    = [r for r in settled if r["result"] == "WIN"]
    losses  = [r for r in settled if r["result"] == "LOSS"]
    pending = [r for r in week_rows if not r["result"]]
    pnl_week = round(sum(r["pnl"] or 0 for r in settled if r["pnl"] is not None), 2)
    win_rate = round(len(wins) / len(settled) * 100, 1) if settled else 0.0

    best_pick = max(wins, key=lambda r: r["pnl"] or 0) if wins else None

    running_total = next(
        (r["running"] for r in reversed(all_rows) if r.get("running") is not None), 0.0
    )

    return {
        "week_start":    week_mon.strftime("%d %b"),
        "week_end":      week_sun.strftime("%d %b %Y"),
        "total_picks":   len(week_rows),
        "wins":          len(wins),
        "losses":        len(losses),
        "pending":       len(pending),
        "win_rate":      win_rate,
        "pnl_week":      pnl_week,
        "best_pick":     best_pick,
        "running_total": running_total or 0.0,
        "pending_picks": pending,
    }


# ── Deduplicate existing Excel data ──────────────────────────────────────────

def cleanup_duplicates() -> int:
    """Remove duplicate pick rows (same date, match, bet_type, pick). Returns rows removed."""
    if not EXCEL_PATH.exists():
        return 0

    wb = load_workbook(EXCEL_PATH)
    ws = wb["Picks"]

    seen: set[tuple] = set()
    rows_to_delete: list[int] = []

    for row in range(2, ws.max_row + 1):
        date_val = ws.cell(row=row, column=1).value
        if date_val is None:
            break
        key_date = date_val.date().isoformat() if isinstance(date_val, datetime) else str(date_val)
        key = (
            key_date,
            str(ws.cell(row=row, column=2).value or ""),
            str(ws.cell(row=row, column=3).value or ""),
            str(ws.cell(row=row, column=4).value or ""),
        )
        if key in seen:
            rows_to_delete.append(row)
        else:
            seen.add(key)

    for row in reversed(rows_to_delete):
        ws.delete_rows(row)

    if rows_to_delete:
        _recalculate_running_total(ws)
        _refresh_summary(wb)
        wb.save(EXCEL_PATH)
        log.info("Removed %d duplicate row(s) from %s", len(rows_to_delete), EXCEL_PATH)
    else:
        log.info("No duplicates found in %s", EXCEL_PATH)

    return len(rows_to_delete)
